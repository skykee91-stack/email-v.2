"""Excel 파일 출력 모듈"""

import logging
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from models.business import Business

logger = logging.getLogger(__name__)


def export_to_excel(businesses: list[Business], filename: str) -> None:
    """업체 목록을 Excel 파일로 저장한다.

    Args:
        businesses: Business 객체 리스트
        filename: 저장할 파일 경로
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "업체 목록"

    # 헤더 스타일
    header_font = Font(bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # 헤더 작성
    headers = [
        "번호", "업체명", "대표전화", "개인번호(010)",
        "이메일", "네이버아이디", "주소", "카테고리", "블로그", "홈페이지",
    ]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # 데이터 행 작성
    for idx, biz in enumerate(businesses, 1):
        row = idx + 1
        values = [
            idx,
            biz.name,
            biz.phone or "",
            biz.personal_phone or "",
            biz.email or "",
            biz.naver_id or "",
            biz.address or "",
            biz.category or "",
            biz.blog_url or "",
            biz.homepage_url or "",
        ]
        for col, value in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = thin_border
            if col == 1:
                cell.alignment = Alignment(horizontal="center")

    # 컬럼 너비 자동 조정
    column_widths = {
        "A": 6, "B": 25, "C": 18, "D": 18,
        "E": 30, "F": 20, "G": 40, "H": 15, "I": 35, "J": 35,
    }
    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width

    # 데이터가 있는 경우 실제 내용에 맞게 너비 재조정
    for col in ws.columns:
        col_letter = col[0].column_letter
        max_length = max(len(str(cell.value or "")) for cell in col)
        adjusted_width = max(max_length + 2, column_widths.get(col_letter, 10))
        ws.column_dimensions[col_letter].width = min(adjusted_width, 50)

    # 필터 추가
    ws.auto_filter.ref = ws.dimensions

    # 첫 번째 행 고정
    ws.freeze_panes = "A2"

    wb.save(filename)
    logger.info(f"Excel 파일 저장 완료: {filename} ({len(businesses)}건)")
