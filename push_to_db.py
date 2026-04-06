"""스크래핑 결과를 Supabase DB에 직접 업로드하는 스크립트

사용법:
    python push_to_db.py 결과파일.xlsx
    python push_to_db.py *.xlsx          (여러 파일 한번에)

이메일이 없는 업체는 건너뛰고, 중복 업체(placeId 또는 이메일)도 건너뜀
"""

import sys
import os
import json
import urllib.request
import urllib.parse

# Supabase 직접 연결 정보
SUPABASE_URL = "https://eejqhprtzjwocepaqzkh.supabase.co"
SUPABASE_KEY = "sb_publishable_jWtxOTobu5HJNHm5PHpnFg_XRMadcyZ"

# 또는 우리 웹 API 사용
WEB_API_URL = "https://email-kappa-teal.vercel.app"


def upload_via_web_api(businesses: list[dict]) -> dict:
    """우리 웹 API를 통해 업체 등록"""
    added = 0
    skipped = 0
    errors = []

    for biz in businesses:
        if not biz.get("email"):
            skipped += 1
            continue

        data = json.dumps(biz, ensure_ascii=False).encode("utf-8")
        req = urllib.request.Request(
            f"{WEB_API_URL}/api/businesses",
            data=data,
            headers={"Content-Type": "application/json; charset=utf-8"},
            method="POST",
        )

        try:
            with urllib.request.urlopen(req, timeout=10) as resp:
                result = json.loads(resp.read())
                if "business" in result:
                    added += 1
                    print(f"  ✅ {biz['name']} ({biz['email']})")
                else:
                    skipped += 1
        except urllib.error.HTTPError as e:
            body = e.read().decode()
            if "이미 등록" in body or "409" in str(e.code):
                skipped += 1
                print(f"  ⏭️ {biz['name']} (이미 등록됨)")
            else:
                errors.append(f"{biz['name']}: {body}")
                print(f"  ❌ {biz['name']}: {body[:100]}")
        except Exception as e:
            errors.append(f"{biz['name']}: {str(e)}")
            print(f"  ❌ {biz['name']}: {e}")

    return {"added": added, "skipped": skipped, "errors": len(errors)}


def read_xlsx(filepath: str) -> list[dict]:
    """Excel 파일에서 업체 데이터 읽기"""
    from openpyxl import load_workbook

    wb = load_workbook(filepath)
    ws = wb.active

    # 헤더 읽기
    headers = []
    for cell in ws[1]:
        headers.append(cell.value or "")

    # 컬럼 매핑 (다양한 헤더명 대응)
    col_map = {}
    for i, h in enumerate(headers):
        h_lower = (h or "").strip().lower().replace(" ", "")
        if h_lower in ["업체명", "name", "상호", "상호명"]:
            col_map["name"] = i
        elif h_lower in ["대표전화", "phone", "전화번호", "전화", "tel"]:
            col_map["phone"] = i
        elif h_lower in ["개인번호(010)", "010번호", "personalphone", "개인번호", "핸드폰"]:
            col_map["personalPhone"] = i
        elif h_lower in ["이메일", "email", "e-mail"]:
            col_map["email"] = i
        elif h_lower in ["네이버아이디", "naverid", "아이디"]:
            col_map["naverId"] = i
        elif h_lower in ["주소", "address"]:
            col_map["address"] = i
        elif h_lower in ["카테고리", "category", "업종"]:
            col_map["category"] = i
        elif h_lower in ["블로그", "blog", "blogurl"]:
            col_map["blogUrl"] = i
        elif h_lower in ["홈페이지", "homepage", "homepageurl", "웹사이트"]:
            col_map["homepageUrl"] = i

    businesses = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[col_map.get("name", 0)]:
            continue

        biz = {"name": str(row[col_map["name"]]).strip()}

        for field, idx in col_map.items():
            if field == "name":
                continue
            value = row[idx] if idx < len(row) else None
            if value:
                biz[field] = str(value).strip()

        businesses.append(biz)

    return businesses


def read_csv(filepath: str) -> list[dict]:
    """CSV 파일에서 업체 데이터 읽기"""
    import csv

    businesses = []
    with open(filepath, "r", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        for row in reader:
            biz = {}
            for key, value in row.items():
                if value and value.strip():
                    # 키 이름 정규화
                    k = key.strip().lower().replace(" ", "")
                    if k in ["업체명", "name", "상호"]:
                        biz["name"] = value.strip()
                    elif k in ["이메일", "email"]:
                        biz["email"] = value.strip()
                    elif k in ["전화번호", "phone", "대표전화"]:
                        biz["phone"] = value.strip()
                    elif k in ["주소", "address"]:
                        biz["address"] = value.strip()
                    elif k in ["업종", "category", "카테고리"]:
                        biz["category"] = value.strip()
                    elif k in ["지역", "region"]:
                        biz["region"] = value.strip()
                    elif k in ["블로그", "blogurl"]:
                        biz["blogUrl"] = value.strip()
                    elif k in ["홈페이지", "homepageurl"]:
                        biz["homepageUrl"] = value.strip()
            if biz.get("name"):
                businesses.append(biz)

    return businesses


def main():
    if len(sys.argv) < 2:
        print("사용법: python push_to_db.py <파일.xlsx 또는 파일.csv>")
        print("예시:")
        print("  python push_to_db.py test_서울_강남_치과_place.xlsx")
        print("  python push_to_db.py data.csv")
        sys.exit(1)

    import glob
    files = []
    for arg in sys.argv[1:]:
        files.extend(glob.glob(arg))

    if not files:
        print(f"파일을 찾을 수 없습니다: {sys.argv[1:]}")
        sys.exit(1)

    total_result = {"added": 0, "skipped": 0, "errors": 0}

    for filepath in files:
        print(f"\n📁 파일: {os.path.basename(filepath)}")

        if filepath.endswith(".xlsx"):
            businesses = read_xlsx(filepath)
        elif filepath.endswith(".csv"):
            businesses = read_csv(filepath)
        else:
            print(f"  지원하지 않는 형식: {filepath}")
            continue

        total = len(businesses)
        with_email = sum(1 for b in businesses if b.get("email"))
        print(f"  총 {total}개 업체, 이메일 보유 {with_email}개 ({with_email/total*100:.1f}%)" if total > 0 else "  빈 파일")

        if total == 0:
            continue

        # 이메일 있는 업체만 업로드
        email_businesses = [b for b in businesses if b.get("email")]
        if not email_businesses:
            print("  이메일 있는 업체가 없어서 건너뜁니다")
            continue

        result = upload_via_web_api(email_businesses)
        print(f"\n  결과: 추가 {result['added']}개, 건너뜀 {result['skipped']}개, 오류 {result['errors']}개")

        total_result["added"] += result["added"]
        total_result["skipped"] += result["skipped"]
        total_result["errors"] += result["errors"]

    print(f"\n{'='*50}")
    print(f"전체 결과: 추가 {total_result['added']}개, 건너뜀 {total_result['skipped']}개, 오류 {total_result['errors']}개")


if __name__ == "__main__":
    main()
